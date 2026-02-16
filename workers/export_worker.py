"""
Export Worker - Handle file exports in background
"""

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from PyQt6.QtCore import pyqtSignal

from workers.base_worker import BaseWorker, TaskPriority


class ExportWorker(BaseWorker):
    """
    Worker for exporting data in background

    Prevents UI freezing during large exports
    """

    # Signals
    export_completed = pyqtSignal(str)  # file_path
    export_progress = pyqtSignal(int)  # progress (0-100)

    def __init__(self, logger: logging.Logger, export_dir: Path):
        super().__init__("ExportWorker", logger)
        self.export_dir = export_dir

        # Ensure export directory exists
        self.export_dir.mkdir(parents=True, exist_ok=True)

        # Connect signals
        self.signals.finished.connect(self._on_export_finished)
        self.signals.progress.connect(self._on_export_progress)

    def export_to_csv(
        self, data: List[Dict[str, Any]], filename: str, headers: List[str]
    ):
        """
        Export data to CSV

        Args:
            data: List of dictionaries
            filename: Output filename
            headers: CSV headers
        """
        task_id = f"export_csv_{datetime.now().timestamp()}"

        self.add_task(
            task_id=task_id,
            func=self._export_csv_task,
            args=(task_id, data, filename, headers),
            priority=TaskPriority.NORMAL,
        )

    def export_to_excel(
        self, data: List[Dict[str, Any]], filename: str, sheet_name: str = "Sheet1"
    ):
        """
        Export data to Excel

        Args:
            data: List of dictionaries
            filename: Output filename
            sheet_name: Excel sheet name
        """
        task_id = f"export_excel_{datetime.now().timestamp()}"

        self.add_task(
            task_id=task_id,
            func=self._export_excel_task,
            args=(task_id, data, filename, sheet_name),
            priority=TaskPriority.NORMAL,
        )

    def _export_csv_task(
        self,
        task_id: str,
        data: List[Dict[str, Any]],
        filename: str,
        headers: List[str],
    ) -> str:
        """
        Export to CSV (runs on worker thread)

        Returns:
            Path to exported file
        """
        # Sanitize filename
        from core.security import InputValidator

        filename = InputValidator.sanitize_filename(filename)

        # Ensure .csv extension
        if not filename.endswith(".csv"):
            filename += ".csv"

        file_path = self.export_dir / filename

        # Write CSV
        total_rows = len(data)
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()

            for i, row in enumerate(data):
                writer.writerow(row)

                # Report progress every 100 rows
                if i % 100 == 0:
                    progress = int((i / total_rows) * 100)
                    self.signals.progress.emit(task_id, progress)

        self.logger.info(f"Exported {total_rows} rows to {file_path}")
        return str(file_path)

    def _export_excel_task(
        self, task_id: str, data: List[Dict[str, Any]], filename: str, sheet_name: str
    ) -> str:
        """
        Export to Excel (runs on worker thread)

        Returns:
            Path to exported file
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")

        # Sanitize filename
        from core.security import InputValidator

        filename = InputValidator.sanitize_filename(filename)

        # Ensure .xlsx extension
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"

        file_path = self.export_dir / filename

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb.save(file_path)
            return str(file_path)

        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Write data
        total_rows = len(data)
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

            # Report progress every 100 rows
            if row_idx % 100 == 0:
                progress = int((row_idx / total_rows) * 100)
                self.signals.progress.emit(task_id, progress)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(file_path)

        self.logger.info(f"Exported {total_rows} rows to {file_path}")
        return str(file_path)

    def _on_export_finished(self, task_id: str, file_path: str):
        """Handle export completion (runs on UI thread)"""
        self.export_completed.emit(file_path)

    def _on_export_progress(self, task_id: str, progress: int):
        """Handle export progress (runs on UI thread)"""
        self.export_progress.emit(progress)
